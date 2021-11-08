
import locale
from pathlib import Path, PosixPath
from typing import List

import numpy as np
import openpyxl as xl
import pandas as pd
from openpyxl.chart import BarChart, Reference, ScatterChart, Series
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrapText=False)
NUM_DIGIT = 2  # 有効桁数
DATE_FORMAT = "%Y-%m-%d（%a）"


class ExcelOutputer:
    def __init__(self):
        pass

    @staticmethod
    def df2excel(df_list: List, sheetname_list: List, output_path: PosixPath):
        with pd.ExcelWriter(output_path) as writer:
            for sheet_name, df in zip(sheetname_list, df_list):
                df.to_excel(writer, sheet_name=sheet_name, index=False)


class ExcelFormater:
    """
    Excelファイルを整形する関数をまとめたクラス

    convert_num_format: 桁を揃える
    adjust_width:  幅を揃える
    align_center_row:   中央揃え
    """

    def __init__(self):
        pass

    @staticmethod
    def convert_num_format(ws: Worksheet, col_name: str, num_digit: int) -> Worksheet:
        """小数点以下の桁数のフォーマットを統一する関数

        Args:
            ws (Worksheet): 対象のエクセルワークシート
            col_name (str): ws内のカラム名
            num_digit (int): 小数点以下何桁を表示するか

        Returns:
            Worksheet: 対象のカラムの小数点以下の桁数を統一したエクセルワークシート
        """
        for wb in ws.worksheets:
            for row in wb:
                for cell in row:
                    col = cell.column
                    if wb.cell(row=1, column=col).value == col_name:
                        cell.number_format = "0." + "0" * num_digit
        return ws

    def convert_format(
        ws: Worksheet, convert_contain_str: str, convert_format: str = "0.00%"
    ) -> Worksheet:
        """
        任意の文字が含まれたカラムを、任意のフォーマットに変換する関数

        Args:
            ws (Worksheet): 対象のエクセルワークシート
            convert_contain_str (str): どのような文字が含まれたカラムを対象とするか
            convert_format (str, optional):どのようなフォーマットにするか（ex. 0.00%、0.00、etc...）
            ※ ref: https://qiita.com/github-nakasho/items/7db0c3bfaad6637a0f40

        Returns:
            Worksheet: 対象のカラムを変換したワークシート
        """
        # カラムリストの取得
        columns_list = [row for i, row in enumerate(ws) if i == 0]
        columns_list = [cell.internal_value for cell in columns_list[0]]

        # 型を変換
        for row in ws:
            for cell in row:
                col = columns_list[cell.column - 1]
                if convert_contain_str in col:
                    cell.number_format = convert_format
        return ws

    @staticmethod
    def adjust_width(ws: Worksheet) -> Worksheet:
        """入力されたシートでカラムごとに最大の文字数に合わせてセル幅を調整する関数

        Args:
            ws (Worksheet): エクセルシート

        Returns:
            Worksheet: セル幅を調整したエクセルシート
        """
        for wb in ws.worksheets:
            for col in wb.columns:
                # デフォルトのセルの幅は8とする
                max_length = 8
                column = col[0].column

                for cell in col:
                    # 日本語は2バイトで、セル内の幅の調整がそのままだとできないので、
                    # バイトにエンコードして幅を調整
                    if len(str(cell.value).encode("shift-jis")) > max_length:
                        max_length = len(str(cell.value).encode("shift-jis"))

                adjusted_width = max_length + 4
                wb.column_dimensions[get_column_letter(column)].width = adjusted_width

        return ws

    @staticmethod
    def align_center_row(ws):
        """row_num行のセルは中央揃えする関数

        Args:
            ws (Worksheet): エクセルシート
            row_num (int): 行番号

        Returns:
            Worksheet: 指定した行を中央揃えしたエクセルシート
        """
        for wb in ws.worksheets:
            for row in wb:
                for cell in row:
                    # if cell.row == row_num:
                    cell.alignment = CENTER_ALIGNMENT
        return ws



class ExcelPlotter:
    def __init__(self):

        self.min_row = 1
        self.max_row = 188  # TODO:動的に決定したい
        self.width = 17.78
        self.height = 5.43

    def create_barchart(self, ws, title, plot_cols_lines, place, y_label, x_label="日付"):
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = title
        chart.y_axis.title = y_label
        chart.x_axis.title = x_label

        # サイズの設定
        chart.width = self.width * 2.5
        chart.height = self.height * 2.5

        # 縦軸のデータをセット
        for wb in ws.worksheets:
            for plot_col in plot_cols_lines:
                data = Reference(
                    wb,
                    min_col=plot_col,
                    min_row=self.min_row,
                    max_row=self.max_row,
                    max_col=plot_col,
                )
                chart.add_data(data, titles_from_data=True)

            # 横軸のデータをセット
            cats = Reference(
                wb,
                min_col=1,
                min_row=self.min_row + 1,
                max_row=self.max_row,
            )
            chart.set_categories(cats)

            chart.shape = 4
            wb.add_chart(chart, place)  # グラフの位置の設定

        return ws

    def plot(
        self,
        ws: Worksheet,
        title: str,
        plot_cols_lines: List,
        y_label: str,
        x_label: str,
        place: str = "Z10",
    ):
        ws = self.create_barchart(
            ws, title, plot_cols_lines, place, y_label, x_label="日付"
        )

        return ws

    def multi_plot(self):
        """
        複数のグラフを同時にプロットする（必要があれば、参考にして改編）
        """
        # プロットのための設定
        place_col = "Z"
        init_row_number = 10
        plot_dict = {"pe": [], "ape": [9, 10], "物単予測誤差": [5, 7], "台数予測誤差": [13, 15]}

        for y_label, plot_cols in plot_dict.items():
            title = f"{y_label}の推移"
            place = f"{place_col}{init_row_number}"
            ws = self.create_barchart(
                ws, title, plot_cols, place, y_label, x_label="日付"
            )
            init_row_number += 30

        return ws
